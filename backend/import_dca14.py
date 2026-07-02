# -*- coding: utf-8 -*-
"""Import lô DCA14 (51 ô) vào cell + mortgage — KHÔNG đụng dữ liệu lô khác.

An toàn (idempotent): xoá riêng dữ liệu cell/mortgage của DCA14 rồi nạp lại,
KHÔNG TRUNCATE (khác import_excel.py cũ vốn xoá sạch mọi lô).

Nguồn:
  - DCA14_tong_theo_o.xlsx  → 51 ô (master cấp ô)
  - DCA14_Data_PLTC.xlsx    → thế chấp cấp lô (đồng nhất: Chủ dự án/Tân Long/Vietinbank)

Geom: nếu có file mapping dca14_geom_mapping.json (mã ô → ranh_thua_id) thì gắn
geom; nếu không, ranh_thua_id=NULL (hiện ở màn Quản lý theo ô, chưa lên bản đồ).

Chạy:
  docker compose cp import_dca14.py api:/app/import_dca14.py
  docker compose cp _excel/... (mount)  # hoặc EXCEL_DIR
  docker compose exec api python import_dca14.py
"""
import os
import re
import json

import openpyxl

from app.db import get_conn

EXCEL_DIR = os.environ.get("EXCEL_DIR", "_excel")
F_TONG = os.environ.get("F_TONG_DCA14", "dca14_tong.xlsx")
F_PL = os.environ.get("F_PL_DCA14", "dca14_pltc.xlsx")
GEOM_MAP = os.environ.get("GEOM_MAP_DCA14", "dca14_geom_mapping.json")

LOT_CODE = "DCA14"

# --- Mapping enum Việt → mã FE (dùng SCHEMA MỚI đã migrate) ---
PLANNING = {
    "Đất ở + dịch vụ thương mại": "residential_commercial",
    "Đất ở chia lô": "residential_lot",
    "Đất nhà vườn": "garden_house",
}
# book_status: schema MỚI (7 mã). "Chưa cấp sổ" → no_book_ineligible.
BOOK = {
    "Chưa cấp sổ": "no_book_ineligible",
    "Chưa cấp sổ - Không đủ điều kiện": "no_book_ineligible",
    "Chưa cấp sổ - Đủ điều kiện": "no_book_eligible",
    "Đang làm thủ tục cấp sổ": "book_in_progress",
    "Đã có sổ - Do CĐT cầm": "book_held_investor",
    "Đã có sổ - Chuyển giao sổ cho chủ sở hữu": "book_transferred",
    "Đã có sổ - Đang giao dịch tài chính / pháp lý": "book_in_transaction",
    "Đã có sổ - Đổi/Tách từ sổ cũ": "book_split",
}
CONSTRUCTION = {"Chưa giao": "not_handed_over", "Đã xây dựng": "built"}


def biz_status(ttkd, ttso):
    if (ttkd or "").strip() == "Chưa bán":
        return "unsold"
    if ttso and "Đã có sổ" in ttso:
        return "sold_red_book"
    return "sold_no_book"


def parse_density(v):
    if v is None:
        return None
    m = re.search(r"([\d,\.]+)", str(v))
    if not m:
        return None
    try:
        return round(float(m.group(1).replace(",", ".")), 2)
    except ValueError:
        return None


def parse_floors(v):
    if v is None:
        return (None, None)
    nums = re.findall(r"\d+", str(v))
    if not nums:
        return (None, None)
    if len(nums) == 1:
        return (int(nums[0]), int(nums[0]))
    return (int(nums[0]), int(nums[1]))


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = re.sub(r"[^\d.\-]", "", str(v).replace(",", ""))
    if not s or s in ("-", "."):
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def load(path):
    return openpyxl.load_workbook(path, data_only=True).active


def main():
    base = EXCEL_DIR
    ws_tong = load(os.path.join(base, F_TONG))
    ws_pl = load(os.path.join(base, F_PL))

    # File tổng: header ở dòng 1, data từ dòng 2. Lấy dòng có Mã ô (idx 2).
    tong = [r for r in ws_tong.iter_rows(min_row=2, values_only=True) if r[2]]
    # File PLTC: header ở dòng 2, data từ dòng 3. Mã ô ở idx 0.
    pl = [r for r in ws_pl.iter_rows(min_row=3, values_only=True) if r[0]]
    print(f"Đọc Excel: {len(tong)} ô, {len(pl)} dòng pháp lý")

    # Geom mapping (mã ô → ranh_thua_id), nếu có.
    geom_map = {}
    geom_path = os.path.join(base, GEOM_MAP)
    if os.path.exists(geom_path):
        with open(geom_path, encoding="utf-8") as f:
            geom_map = json.load(f)
        print(f"Geom mapping: {len(geom_map)} ô có ranh_thua_id")
    else:
        print("Chưa có geom mapping → ranh_thua_id=NULL (chưa lên bản đồ)")

    with get_conn() as conn, conn.cursor() as cur:
        # subdivision DCA14 đã seed sẵn → dùng lại.
        cur.execute("SELECT id FROM subdivision WHERE lot_code=%s", (LOT_CODE,))
        row = cur.fetchone()
        if not row:
            raise SystemExit(f"Không tìm thấy subdivision {LOT_CODE} — cần seed trước.")
        sub_id = row[0]

        # Reset RIÊNG dữ liệu DCA14 (idempotent, không đụng lô khác).
        cur.execute(
            "DELETE FROM mortgage WHERE subdivision_id=%s", (sub_id,)
        )
        cur.execute(
            "DELETE FROM cell WHERE subdivision_id=%s", (sub_id,)
        )

        # Chủ sở hữu / người đứng tên: file TONG (cột 10) TRỐNG cho DCA14 →
        # lấy từ file PLTC cột "Người đứng tên" (idx 5, VD "Chủ dự án").
        # Lô chưa bán → chủ đầu tư ("Chủ dự án") vẫn là chủ sở hữu hiện tại.
        owner_by_cell = {}
        for r in pl:
            code = str(r[0]).strip()
            if r[5]:
                owner_by_cell[code] = str(r[5]).strip()

        # CELL — 51 ô. Cột TONG: 0 Lô,1 Ô,2 Mã ô,3 ĐC,4 Tọa độ,5 DT,6 Loại đất,
        # 7 TT sổ,8 Số sổ,9 TT KD,10 Chủ SH,11 HĐ,12 TGGD,13 THA,14 TT XD,
        # 15 Mật độ,16 Số tầng
        n_cell = n_geom = 0
        for r in tong:
            cell_code = str(r[2]).strip()
            # Ưu tiên chủ SH từ file tổng (r[10]); nếu trống → người đứng tên (PLTC).
            owner = r[10] if r[10] else owner_by_cell.get(cell_code)
            biz = biz_status(r[9], r[7])
            book = BOOK.get((r[7] or "").strip())
            constr = CONSTRUCTION.get((r[14] or "").strip())
            dmin, dmax = parse_floors(r[16])
            raw = {
                "loai_dat": r[6], "tinh_trang_so": r[7], "tinh_trang_kd": r[9],
                "chu_so_huu": r[10], "tinh_trang_xd": r[14],
                "mat_do": r[15], "so_tang": r[16],
            }
            ranh_id = geom_map.get(cell_code)

            if ranh_id is not None:
                cur.execute(
                    """INSERT INTO cell
                       (cell_code, subdivision_id, cell_no, ranh_thua_id, centroid,
                        owner_name, address, area, planning_type, business_status,
                        book_status, book_no, construction_status,
                        build_density, build_floor_min, build_floor_max,
                        collateral_status, raw_excel)
                       VALUES (%s,%s,%s,%s,
                               (SELECT ST_Centroid(geom) FROM ranh_thua WHERE id=%s),
                               %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (cell_code, sub_id, r[1], ranh_id, ranh_id,
                     owner, r[3], to_num(r[5]), PLANNING.get(r[6]), biz,
                     book, r[8], constr,
                     parse_density(r[15]), dmin, dmax,
                     "mortgage_bank", json.dumps(raw, ensure_ascii=False)),
                )
                n_geom += 1
            else:
                cur.execute(
                    """INSERT INTO cell
                       (cell_code, subdivision_id, cell_no,
                        owner_name, address, area, planning_type, business_status,
                        book_status, book_no, construction_status,
                        build_density, build_floor_min, build_floor_max,
                        collateral_status, raw_excel)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (cell_code, sub_id, r[1],
                     owner, r[3], to_num(r[5]), PLANNING.get(r[6]), biz,
                     book, r[8], constr,
                     parse_density(r[15]), dmin, dmax,
                     "mortgage_bank", json.dumps(raw, ensure_ascii=False)),
                )
            n_cell += 1

        # MORTGAGE — cấp lô (PLTC đồng nhất). Cột PL: 0 Mã ô,1 Mã lô,2 DT,
        # 3 TT pháp lý,4 Số sổ,5 Người đứng tên,6 Bên vay,7 Tổ chức TC,
        # 8 Giá trị,9 Loại,10 Mục đích,11 Ghi chú
        # 3 vai trò riêng: holder_name="Người đứng tên" (idx5, "Chủ dự án"),
        # borrower_name="Bên đứng tên vay" (idx6, "Tân Long"),
        # lender_name="Tổ chức TC" (idx7, "Vietinbank").
        if pl:
            r0 = pl[0]
            cur.execute(
                """INSERT INTO mortgage
                   (subdivision_id, state, lender_name, borrower_name,
                    holder_name, mortgage_type, purpose)
                   VALUES (%s,'mortgaged',%s,%s,%s,%s,%s)""",
                (sub_id, r0[7], r0[6], r0[5], r0[9],
                 r0[10] if len(r0) > 10 else None),
            )

        conn.commit()

        # VERIFY
        cur.execute("SELECT count(*) FROM cell WHERE subdivision_id=%s", (sub_id,))
        print(f"  cell DCA14        : {cur.fetchone()[0]}")
        cur.execute(
            "SELECT count(*) FROM cell WHERE subdivision_id=%s "
            "AND ranh_thua_id IS NOT NULL", (sub_id,))
        print(f"  cell có geom      : {cur.fetchone()[0]}")
        cur.execute("SELECT count(*) FROM mortgage WHERE subdivision_id=%s", (sub_id,))
        print(f"  mortgage DCA14    : {cur.fetchone()[0]}")
    print(f"\nXong. {n_cell} ô, {n_geom} gắn geom.")


if __name__ == "__main__":
    main()
