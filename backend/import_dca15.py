# -*- coding: utf-8 -*-
"""Import lô DCA15 vào cell + cell_owner + contract + payment — KHÔNG đụng lô khác.

Chỉ import các Ô ĐÃ BÁN (có trong file HDGD, 23 ô). File 'tổng theo ô' làm gốc
(diện tích, sổ, XD...); HDGD bổ sung tài chính (giá trị, thanh toán, thuế, HĐ).

Đặc biệt:
  - Nhiều chủ sở hữu: Excel ghi "1.Tên A\\n2.Tên B" trong 1 ô → tách thành
    nhiều dòng cell_owner (theo ordinal).
  - tax_bearer: "Có VAT" → mã 'has_vat' (đã mở rộng CHECK constraint).

Geom: nếu có dca15_geom_mapping.json thì gắn; nếu không, ranh_thua_id=NULL.

An toàn (idempotent): DELETE riêng dữ liệu DCA15 rồi nạp lại, KHÔNG TRUNCATE.

Chạy:
  docker cp import_dca15.py backend-api-1:/app/
  docker cp _excel/... (mount)
  docker exec -e ... backend-api-1 python /app/import_dca15.py
"""
import os
import re
import json
import datetime as dt

import openpyxl

from app.db import get_conn

EXCEL_DIR = os.environ.get("EXCEL_DIR", "_excel")
F_TONG = os.environ.get("F_TONG_DCA15", "dca15_tong.xlsx")
F_HD = os.environ.get("F_HD_DCA15", "dca15_hdgd.xlsx")
GEOM_MAP = os.environ.get("GEOM_MAP_DCA15", "dca15_geom_mapping.json")

LOT_CODE = "DCA15"

PLANNING = {
    "Đất ở + dịch vụ thương mại": "residential_commercial",
    "Đất ở chia lô": "residential_lot",
    "Đất nhà vườn": "garden_house",
}
BOOK = {
    "Chưa cấp sổ": "no_book_ineligible",
    "Chưa cấp sổ - Không đủ điều kiện": "no_book_ineligible",
    "Chưa cấp sổ - Đủ điều kiện": "no_book_eligible",
    "Đang làm thủ tục cấp sổ": "book_in_progress",
    "Đã có sổ - Do CĐT cầm": "book_held_investor",
    "Đã có sổ - Chuyển giao sổ cho chủ sở hữu": "book_transferred",
    "Đã có sổ - Đang giao dịch tài chính / pháp lý": "book_in_transaction",
    "Đã có sổ - Đổi/Tách từ sổ cũ": "book_split",
}
CONSTRUCTION = {"Chưa giao": "not_handed_over", "Đã xây dựng": "built"}
TAX_BEARER = {
    "Khách hàng": "customer", "Chủ đầu tư": "investor",
    "Bên A": "investor", "Bên B": "customer", "Có VAT": "has_vat",
}


def biz_status(ttkd, ttso):
    s = (ttkd or "").strip()
    if s == "Chưa bán":
        return "unsold"
    if ttso and "Đã có sổ" in ttso:
        return "sold_red_book"
    return "sold_no_book"


def pay_status(tttt):
    if not tttt:
        return None
    s = tttt.strip()
    if s == "Đã thanh toán":
        return "paid_full"
    if "một phần" in s.lower() or s == "Thanh toán một phần":
        return "partial"
    return None


def parse_density(v):
    if v is None:
        return None
    m = re.search(r"([\d,\.]+)", str(v))
    if not m:
        return None
    try:
        return round(float(m.group(1).replace(",", ".")), 2)
    except ValueError:
        return None


def parse_floors(v):
    if v is None:
        return (None, None)
    nums = re.findall(r"\d+", str(v))
    if not nums:
        return (None, None)
    if len(nums) == 1:
        return (int(nums[0]), int(nums[0]))
    return (int(nums[0]), int(nums[1]))


def to_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def to_num(v):
    """Chuẩn hoá số kiểu VN → int. Xử lý:
      '2.523.000.000' (chấm ngăn nghìn) → 2523000000
      '750,000,000'   (phẩy ngăn nghìn) → 750000000
      '112,5' / '87,5' (PHẨY thập phân) → 113 / 88 (làm tròn)
      '87.00' (chấm thập phân) / 750000000 / int."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = str(v).strip()
    # Phẩy thập phân kiểu VN: đúng 1 dấu phẩy + theo sau 1-2 chữ số ở CUỐI.
    if re.fullmatch(r"-?\d+,\d{1,2}", s):
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")  # còn lại: phẩy = ngăn cách nghìn → bỏ
    # Chấm ngăn nghìn: nhiều dấu chấm HOẶC chấm theo sau đúng 3 chữ số.
    if s.count(".") > 1 or re.search(r"\.\d{3}(\D|$)", s):
        s = s.replace(".", "")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in ("-", "."):
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def parse_owners(raw):
    """Tách chuỗi chủ sở hữu → list.
    "1.Tên A\\n2.Tên B" → [Tên A, Tên B]; "Tên" → [Tên]."""
    if not raw:
        return []
    s = str(raw).strip()
    # Tách theo xuống dòng HOẶC pattern "số." / "số)" đầu mỗi mục.
    items = re.split(r"\n|\s*\d+\s*[.)]\s*", s)
    items = [x.strip() for x in items if x and x.strip()]
    return items if items else [s]


def to_area(v):
    """Diện tích → float (GIỮ thập phân, VD '112,5' → 112.5). cell.area là
    numeric(12,2) nên không làm tròn về int như to_num."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if re.fullmatch(r"-?\d+,\d{1,2}", s):
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    if s.count(".") > 1:
        s = s.replace(".", "")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def load(path):
    return openpyxl.load_workbook(path, data_only=True).active


def main():
    base = EXCEL_DIR
    ws_tong = load(os.path.join(base, F_TONG))
    ws_hd = load(os.path.join(base, F_HD))

    # File tổng: header dòng 1, data từ dòng 2. Mã ô idx 2.
    tong = {str(r[2]).strip(): r for r in ws_tong.iter_rows(min_row=2, values_only=True) if r[2]}
    # HDGD: header dòng 2, data từ dòng 3. Mã ô idx 2.
    hd_rows = [r for r in ws_hd.iter_rows(min_row=3, values_only=True) if r[2]]
    # Chỉ import ô CÓ trong HDGD (đã bán). Gom HDGD theo mã ô — GIỮ TẤT CẢ dòng
    # (1 ô có thể nhiều dòng = nhiều ĐỢT thanh toán = lộ trình chi tiết).
    hd_all_by_cell = {}
    for r in hd_rows:
        code = str(r[2]).strip()
        hd_all_by_cell.setdefault(code, []).append(r)
    # Dòng đầu mỗi ô làm đại diện cho cell + contract (giá trị HĐ, chủ, thuế...).
    hd_by_cell = {code: rows[0] for code, rows in hd_all_by_cell.items()}
    cells_to_import = list(hd_by_cell.keys())
    print(f"Đọc Excel: tổng {len(tong)} ô, HDGD {len(cells_to_import)} ô đã bán")

    geom_map = {}
    geom_path = os.path.join(base, GEOM_MAP)
    if os.path.exists(geom_path):
        with open(geom_path, encoding="utf-8") as f:
            geom_map = json.load(f)
        print(f"Geom mapping: {len(geom_map)} ô")
    else:
        print("Chưa có geom mapping → ranh_thua_id=NULL")

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT id FROM subdivision WHERE lot_code=%s", (LOT_CODE,))
        row = cur.fetchone()
        if not row:
            raise SystemExit(f"Không tìm thấy subdivision {LOT_CODE}")
        sub_id = row[0]

        # Reset riêng DCA15 (cell CASCADE → cell_owner/payment/contract_cell tự xoá).
        cur.execute(
            "DELETE FROM contract WHERE subdivision_id=%s", (sub_id,))
        cur.execute("DELETE FROM cell WHERE subdivision_id=%s", (sub_id,))

        n_cell = n_geom = n_owner = 0
        cell_id_by_code = {}
        for code in cells_to_import:
            t = tong.get(code)
            h = hd_by_cell[code]
            if not t:
                print(f"  ⚠ {code} có trong HDGD nhưng KHÔNG có trong file tổng — bỏ qua")
                continue
            # t cols: 0 Lô,1 Ô,2 Mã ô,3 ĐC,4 Tọa độ,5 DT,6 Loại,7 TTsổ,8 Số sổ,
            #         9 TTKD,10 Chủ SH,11 HĐ,12 TGGD,13 THA,14 TTXD,15 Mật độ,16 Số tầng
            biz = biz_status(t[9], t[7])
            book = BOOK.get((t[7] or "").strip())
            constr = CONSTRUCTION.get((t[14] or "").strip())
            dmin, dmax = parse_floors(t[16])
            # h cols: 0 Lô,1 Ô,2 Mã ô,3 DT,4 Chủ,5 ĐC,6 Số sổ,7 HĐclean,8 HĐorig,
            #         9 TGGD,10 GiáHĐ,11 Thuế,12 Bên thuế,13 Đơn giá,14 TTthanh toán,
            #         15 Đã TT,16 Còn lại,17 Ghi chú
            value = to_num(h[10])
            # paid_value = TỔNG các đợt thanh toán (ô nhiều đợt → cộng hết).
            paid = sum(
                (to_num(r[15]) or 0) for r in hd_all_by_cell[code]
            ) or None
            # remaining: lấy "còn lại" ở dòng CUỐI (phản ánh sau đợt mới nhất);
            # nếu trống → value - paid.
            remaining = to_num(hd_all_by_cell[code][-1][16])
            if remaining is None and value is not None and paid is not None:
                remaining = value - paid
            pay = pay_status(h[14])
            address = t[3] or h[5]
            raw = {
                "loai_dat": t[6], "tinh_trang_so": t[7], "tinh_trang_kd": t[9],
                "tinh_trang_xd": t[14], "mat_do": t[15], "so_tang": t[16],
            }
            ranh_id = geom_map.get(code)

            common = (
                address, to_area(t[5]), PLANNING.get(t[6]), biz,
                book, t[8], constr,
                parse_density(t[15]), dmin, dmax,
                value, paid, remaining, pay,
                "none", json.dumps(raw, ensure_ascii=False),
            )
            if ranh_id is not None:
                cur.execute(
                    """INSERT INTO cell
                       (cell_code, subdivision_id, cell_no, ranh_thua_id, centroid,
                        address, area, planning_type, business_status,
                        book_status, book_no, construction_status,
                        build_density, build_floor_min, build_floor_max,
                        value, paid_value, remaining_value, payment_status,
                        collateral_status, raw_excel)
                       VALUES (%s,%s,%s,%s,
                               (SELECT ST_Centroid(geom) FROM ranh_thua WHERE id=%s),
                               %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       RETURNING id""",
                    (code, sub_id, t[1], ranh_id, ranh_id) + common,
                )
            else:
                cur.execute(
                    """INSERT INTO cell
                       (cell_code, subdivision_id, cell_no,
                        address, area, planning_type, business_status,
                        book_status, book_no, construction_status,
                        build_density, build_floor_min, build_floor_max,
                        value, paid_value, remaining_value, payment_status,
                        collateral_status, raw_excel)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       RETURNING id""",
                    (code, sub_id, t[1]) + common,
                )
            cid = cur.fetchone()[0]
            cell_id_by_code[code] = cid
            n_cell += 1
            if ranh_id is not None:
                n_geom += 1

            # cell_owner — tách nhiều chủ.
            owners = parse_owners(t[10] or h[4])
            for i, name in enumerate(owners, start=1):
                cur.execute(
                    "INSERT INTO cell_owner (cell_id, owner_name, ordinal) "
                    "VALUES (%s,%s,%s)",
                    (cid, name, i),
                )
                n_owner += 1

        # CONTRACT + contract_cell + payment (từ HDGD, mỗi ô 1 dòng đại diện).
        n_contract = n_payment = 0
        contract_by_code = {}
        for code, h in hd_by_cell.items():
            if code not in cell_id_by_code:
                continue
            code_clean = str(h[7]).strip() if h[7] else None
            customer = h[4]
            if code_clean and code_clean not in contract_by_code:
                cur.execute(
                    """INSERT INTO contract
                       (subdivision_id, code_clean, code_original, customer_name,
                        sign_date, tax_amount, tax_bearer, unit_price, note)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (sub_id, code_clean,
                     str(h[8]) if h[8] is not None else None,
                     str(customer) if customer else None,
                     to_date(h[9]), to_num(h[11]),
                     TAX_BEARER.get((str(h[12]) or "").strip()),
                     to_num(h[13]), h[17]),
                )
                contract_by_code[code_clean] = cur.fetchone()[0]
                n_contract += 1
            if code_clean:
                cur.execute(
                    "INSERT INTO contract_cell (contract_id, cell_id) VALUES (%s,%s) "
                    "ON CONFLICT DO NOTHING",
                    (contract_by_code[code_clean], cell_id_by_code[code]),
                )
            # PAYMENT — LỘ TRÌNH: mỗi DÒNG Excel của ô = 1 đợt thanh toán (đúng
            # ngày + số tiền từng đợt). Ô có 2 dòng → 2 payment.
            for pay_row in hd_all_by_cell[code]:
                paid = to_num(pay_row[15])
                if paid:
                    cur.execute(
                        "INSERT INTO payment (cell_id, amount, paid_date, source, note) "
                        "VALUES (%s,%s,%s,'excel_aggregate',%s)",
                        (cell_id_by_code[code], paid, to_date(pay_row[9]),
                         "Thanh toán theo hợp đồng"),
                    )
                    n_payment += 1

        conn.commit()

        # VERIFY
        cur.execute("SELECT count(*) FROM cell WHERE subdivision_id=%s", (sub_id,))
        print(f"  cell            : {cur.fetchone()[0]}")
        cur.execute(
            "SELECT count(*) FROM cell_owner co JOIN cell c ON c.id=co.cell_id "
            "WHERE c.subdivision_id=%s", (sub_id,))
        print(f"  cell_owner      : {cur.fetchone()[0]}")
        cur.execute("SELECT count(*) FROM contract WHERE subdivision_id=%s", (sub_id,))
        print(f"  contract        : {cur.fetchone()[0]}")
    print(f"\nXong. {n_cell} ô, {n_geom} geom, {n_owner} chủ, "
          f"{n_contract} HĐ, {n_payment} thanh toán.")


if __name__ == "__main__":
    main()
