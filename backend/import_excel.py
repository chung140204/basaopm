# -*- coding: utf-8 -*-
"""Import 4 file Excel nghiệp vụ vào schema_v2 (cell/contract/payment/mortgage/legal_event).

Nguồn:
  - Data tổng theo ô.xlsx              → master cấp ô (59 ô, 2 lô)
  - Data theo HĐ giao dịch.xlsx        → 17 ô đã bán (DCB02), 9 HĐ (1 HĐ↔nhiều ô)
  - Data theo tình trạng pháp lý tài chính.xlsx → thế chấp (đồng nhất theo lô)
  - Data theo tình trạng thi hành án.xlsx       → timeline THA (cấp lô DCB02)

Đặc điểm:
  - Excel là NGUỒN nghiệp vụ chính. cell_code ('DCB02-01') = business key.
  - DCB02: map geom thật qua ID_BY_OTO (bảng mapping đã căn tay, nhúng bên dưới).
  - DCB09: chưa có mapping geom → ranh_thua_id=NULL, centroid=NULL.
  - Idempotent: xoá sạch dữ liệu nghiệp vụ (TRUNCATE) rồi nạp lại. KHÔNG đụng ranh_thua/lo.

Chạy:
  docker compose cp import_excel.py api:/app/import_excel.py
  docker compose cp "../../Data tổng theo ô.xlsx" api:/app/_excel/...   (hoặc mount)
  docker compose exec api python import_excel.py
Hoặc local có DATABASE_URL + 4 file Excel trong EXCEL_DIR.
"""
import os
import re
import json
import datetime as dt

import openpyxl

from app.db import get_conn

# Thư mục chứa 4 file Excel. Mặc định /app/_excel (mount), fallback ../../ (local).
EXCEL_DIR = os.environ.get("EXCEL_DIR", "_excel")
# Tên file ASCII (copy từ bản gốc tiếng Việt — docker cp lỗi với tên Unicode).
# Bản gốc: "Data tổng theo ô.xlsx" / "Data theo HĐ giao dịch.xlsx" /
#          "Data theo tình trạng pháp lý tài chính.xlsx" / "...thi hành án.xlsx"
F_TONG = os.environ.get("F_TONG", "tong.xlsx")
F_HD = os.environ.get("F_HD", "hd.xlsx")
F_PL = os.environ.get("F_PL", "phaply.xlsx")
F_THA = os.environ.get("F_THA", "tha.xlsx")

# --- Geom mapping DCB02: mã ô (STT) → ranh_thua.id (căn tay theo centroid + diện tích)
DCB02_ID_BY_OTO = {
    1: 1339, 2: 1290, 3: 1291, 4: 1292, 5: 1293, 6: 1294, 7: 1340,
    8: 1323, 9: 1324, 10: 1325, 11: 1326,
    12: 1327, 13: 1328, 14: 1329, 15: 1330,
    16: 1335, 17: 1336, 18: 1337, 19: 1338,
    20: 1341, 21: 1342, 22: 1343, 23: 1344,
    24: 1334, 25: 1333, 26: 1332, 27: 1331,
}

# =========================== MAPPING enum Việt → mã FE ===========================
PLANNING = {
    "Đất ở + dịch vụ thương mại": "residential_commercial",
    "Đất ở chia lô": "residential_lot",
    "Đất nhà vườn": "garden_house",
}
BOOK = {
    "Chưa cấp sổ": "none",
    "Đã có sổ - Chuyển giao sổ cho chủ sở hữu": "issued_transferred",
    "Đã có sổ - Đang giao dịch tài chính / pháp lý": "issued_in_progress",
}
CONSTRUCTION = {"Chưa giao": "not_handed_over", "Đã xây dựng": "built"}
TAX_BEARER = {"Khách hàng": "customer", "Chủ đầu tư": "investor",
              "Bên A": "investor", "Bên B": "customer"}


def biz_status(ttkd, ttso):
    """Chưa bán→unsold; Đã bán + (đã có sổ→sold_red_book | chưa sổ→sold_no_book)."""
    if (ttkd or "").strip() == "Chưa bán":
        return "unsold"
    if ttso and "Đã có sổ" in ttso:
        return "sold_red_book"
    return "sold_no_book"


def pay_status(tttt):
    if not tttt:
        return None
    s = tttt.strip()
    if s == "Đã thanh toán":
        return "paid_full"
    if s == "Thanh toán một phần":
        return "partial"
    return None


def parse_density(v):
    """'72,8%' → 72.80 ; None nếu không parse được."""
    if v is None:
        return None
    m = re.search(r"([\d,\.]+)", str(v))
    if not m:
        return None
    try:
        return round(float(m.group(1).replace(",", ".")), 2)
    except ValueError:
        return None


def parse_floors(v):
    """'2 -5 tầng' → (2,5) ; '3 tầng' → (3,3)."""
    if v is None:
        return (None, None)
    nums = re.findall(r"\d+", str(v))
    if not nums:
        return (None, None)
    if len(nums) == 1:
        return (int(nums[0]), int(nums[0]))
    return (int(nums[0]), int(nums[1]))


def to_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def to_num(v):
    """'750,000,000' / 750000000 / '87.00' → int/None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = re.sub(r"[^\d.\-]", "", str(v).replace(",", ""))
    if not s or s in ("-", "."):
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def load(path):
    return openpyxl.load_workbook(path, data_only=True).active


def main():
    base = EXCEL_DIR
    ws_tong = load(os.path.join(base, F_TONG))
    ws_hd = load(os.path.join(base, F_HD))
    ws_pl = load(os.path.join(base, F_PL))
    ws_tha = load(os.path.join(base, F_THA))

    tong = [r for r in ws_tong.iter_rows(min_row=3, values_only=True) if r[2]]
    hd = [r for r in ws_hd.iter_rows(min_row=3, values_only=True) if r[2]]
    pl = [r for r in ws_pl.iter_rows(min_row=3, values_only=True) if r[0]]
    tha = [r for r in ws_tha.iter_rows(min_row=3, values_only=True) if r[0]]
    print(f"Đọc Excel: {len(tong)} ô, {len(hd)} HĐ-dòng, {len(pl)} pháp lý, {len(tha)} THA")

    with get_conn() as conn, conn.cursor() as cur:
        # 0) Reset sạch dữ liệu nghiệp vụ (idempotent). KHÔNG đụng ranh_thua/lo.
        cur.execute(
            "TRUNCATE document, legal_event, mortgage, payment, contract_cell, "
            "contract, cell, subdivision RESTART IDENTITY CASCADE"
        )

        # 1) SUBDIVISION (2 lô). lo_layer_id = layer của geom; lo cụ thể không
        #    có loCode='DCB09' nên để layer_id chung; DCB02 geom ở lo id=33.
        lots = {}
        # zone: phân khu của lô (DCB02/DCB09 đều thuộc Khu B).
        ZONE_BY_LOT = {"DCB02": "khu-b", "DCB09": "khu-b"}
        for code in ["DCB02", "DCB09"]:
            cur.execute(
                "INSERT INTO subdivision (lot_code, name, zone, lo_layer_id) "
                "VALUES (%s,%s,%s,%s) RETURNING id",
                (code, f"Lô {code}", ZONE_BY_LOT.get(code), "truonglinh-chialo"),
            )
            lots[code] = cur.fetchone()[0]

        # 2) Build map mã ô → dòng HĐ (để lấy customer/sign_date/tax/unit_price + tiền)
        hd_by_cell = {r[2]: r for r in hd}     # Mã ô → dòng HĐ
        # cols HD: 0 Lô,1 Ô,2 Mã ô,3 DT,4 Chủ,5 Địa chỉ,6 Số sổ,7 HĐ clean,
        #          8 HĐ orig,9 TGGD,10 Giá trị,11 Thuế,12 Bên thuế,13 Đơn giá,
        #          14 TT thanh toán,15 Đã TT,16 Còn lại,17 Ghi chú

        # 3) CELL (59 ô từ file tổng)
        # cols TONG: 0 Lô,1 Ô,2 Mã ô,3 ĐC,4 Tọa độ,5 DT,6 Loại đất,7 TT sổ,
        #            8 Số sổ,9 TT KD,10 Chủ SH,11 HĐ clean,12 TGGD,13 TT THA,
        #            14 TT XD,15 Mật độ,16 Số tầng
        cell_id_by_code = {}
        n_geom = 0
        for r in tong:
            lot_code = r[0]
            cell_code = r[2]
            stt = to_num(r[1])
            sub_id = lots[lot_code]

            hdr = hd_by_cell.get(cell_code)
            address = hdr[5] if hdr else None

            # geom: chỉ DCB02 có mapping
            ranh_id, centroid_sql, centroid_param = None, "NULL", None
            if lot_code == "DCB02" and stt in DCB02_ID_BY_OTO:
                ranh_id = DCB02_ID_BY_OTO[stt]

            biz = biz_status(r[9], r[7])
            book = BOOK.get((r[7] or "").strip())
            constr = CONSTRUCTION.get((r[14] or "").strip())
            dmin, dmax = parse_floors(r[16])
            tttt = hdr[14] if hdr else None
            pay = pay_status(tttt)

            value = to_num(hdr[10]) if hdr else None
            paid = to_num(hdr[15]) if hdr else None
            remaining = to_num(hdr[16]) if hdr else None

            raw = {
                "loai_dat": r[6], "tinh_trang_so": r[7], "tinh_trang_kd": r[9],
                "chu_so_huu": r[10], "tinh_trang_xd": r[14],
                "mat_do": r[15], "so_tang": r[16],
            }

            # geom centroid: lấy từ ranh_thua nếu có map
            if ranh_id is not None:
                cur.execute(
                    """INSERT INTO cell
                       (cell_code, subdivision_id, cell_no, ranh_thua_id, centroid,
                        owner_name, address, area, planning_type, business_status,
                        book_status, book_no, construction_status,
                        build_density, build_floor_min, build_floor_max,
                        value, paid_value, remaining_value, payment_status,
                        collateral_status, raw_excel)
                       VALUES (%s,%s,%s,%s,
                               (SELECT ST_Centroid(geom) FROM ranh_thua WHERE id=%s),
                               %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       RETURNING id""",
                    (cell_code, sub_id, r[1], ranh_id, ranh_id,
                     r[10], address, to_num(r[5]), PLANNING.get(r[6]), biz,
                     book, r[8], constr,
                     parse_density(r[15]), dmin, dmax,
                     value, paid, remaining, pay,
                     "mortgage_bank", json.dumps(raw, ensure_ascii=False)),
                )
                n_geom += 1
            else:
                cur.execute(
                    """INSERT INTO cell
                       (cell_code, subdivision_id, cell_no,
                        owner_name, address, area, planning_type, business_status,
                        book_status, book_no, construction_status,
                        build_density, build_floor_min, build_floor_max,
                        value, paid_value, remaining_value, payment_status,
                        collateral_status, raw_excel)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       RETURNING id""",
                    (cell_code, sub_id, r[1],
                     r[10], address, to_num(r[5]), PLANNING.get(r[6]), biz,
                     book, r[8], constr,
                     parse_density(r[15]), dmin, dmax,
                     value, paid, remaining, pay,
                     "mortgage_bank", json.dumps(raw, ensure_ascii=False)),
                )
            cell_id_by_code[cell_code] = cur.fetchone()[0]

        # 4) CONTRACT + CONTRACT_CELL (1 HĐ ↔ nhiều ô). Khoá HĐ = (lô, code clean).
        #    Tiền đã ghi vào cell ở bước 3 → contract chỉ giữ metadata.
        contract_id_by_key = {}
        for r in hd:
            lot_code = r[0]
            cell_code = r[2]
            code_clean_full = r[7]   # 'PalmCity-DCB02-12/15'
            code_orig = r[8]
            # Giữ NGUYÊN mã HĐ đầy đủ từ Excel ('PalmCity-DCB02-12/15') — trước
            # đây regex cắt còn '12/15' làm mất tiền tố lô/dự án khi hiển thị.
            code_clean = str(code_clean_full).strip() if code_clean_full else None
            key = (lot_code, code_clean)
            if key not in contract_id_by_key:
                cur.execute(
                    """INSERT INTO contract
                       (subdivision_id, code_clean, code_original, customer_name,
                        sign_date, tax_amount, tax_bearer, unit_price, note)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (lots[lot_code], code_clean,
                     str(code_orig) if code_orig is not None else None,
                     r[4], to_date(r[9]), to_num(r[11]),
                     TAX_BEARER.get((str(r[12]) or "").strip()),
                     to_num(r[13]), r[17]),
                )
                contract_id_by_key[key] = cur.fetchone()[0]
            cid = contract_id_by_key[key]
            if cell_code in cell_id_by_code:
                cur.execute(
                    "INSERT INTO contract_cell (contract_id, cell_id) VALUES (%s,%s) "
                    "ON CONFLICT DO NOTHING",
                    (cid, cell_id_by_code[cell_code]),
                )

            # 5) PAYMENT per-ô (từ 'Giá trị đã thanh toán', source=excel_aggregate)
            paid = to_num(r[15])
            if paid and cell_code in cell_id_by_code:
                cur.execute(
                    "INSERT INTO payment (cell_id, amount, paid_date, source, note) "
                    "VALUES (%s,%s,%s,'excel_aggregate',%s)",
                    (cell_id_by_code[cell_code], paid, to_date(r[9]),
                     "Đã thanh toán (tổng từ Excel HĐ)"),
                )

        # 6) MORTGAGE — cấp lô (File 3: đồng nhất theo lô). 1 bản ghi/lô.
        # cols PL: 0 Mã ô,1 Mã lô,2 DT,3 TT pháp lý,4 Số sổ,5 Người đứng tên,
        #          6 Bên vay,7 Tổ chức TC,8 Giá trị,9 Loại,10 Mục đích
        lot_mortgage = {}   # lot_code → (lender, borrower, type, purpose)
        for r in pl:
            lot_code = r[1]
            if lot_code not in lot_mortgage:
                lot_mortgage[lot_code] = (r[7], r[6], r[9],
                                          r[10] if len(r) > 10 else None)
        for lot_code, (lender, borrower, mtype, purpose) in lot_mortgage.items():
            if lot_code not in lots:
                continue
            cur.execute(
                """INSERT INTO mortgage
                   (subdivision_id, state, lender_name, borrower_name,
                    mortgage_type, purpose)
                   VALUES (%s,'mortgaged',%s,%s,%s,%s)""",
                (lots[lot_code], lender, borrower, mtype, purpose),
            )

        # 7) LEGAL_EVENT — THA cấp lô (File 4: 27 dòng DCB02 cùng nội dung → 1 sự kiện lô)
        # cols THA: 0 Mã ô,1 Mã lô,2 TT pháp lý,3 Số sổ,4 Người đứng tên,
        #           5 Thời gian cập nhật,6 Nội dung,7 Cơ quan THA,8 Tổ chức TC,9 Giá trị
        tha_by_lot = {}
        for r in tha:
            lot_code = r[1]
            if lot_code not in tha_by_lot:
                tha_by_lot[lot_code] = r
        for lot_code, r in tha_by_lot.items():
            if lot_code not in lots:
                continue
            cur.execute(
                """INSERT INTO legal_event
                   (subdivision_id, event_kind, event_date, legal_status,
                    change_content, enforce_agency, holder_name, finance_org,
                    collateral_value)
                   VALUES (%s,'enforcement',%s,%s,%s,%s,%s,%s,%s)""",
                (lots[lot_code], to_date(r[5]), r[2], r[6], r[7], r[4], r[8],
                 to_num(r[9])),
            )

        conn.commit()

        # --- VERIFY ---
        for t in ["subdivision", "cell", "contract", "contract_cell",
                  "payment", "mortgage", "legal_event"]:
            cur.execute(f"SELECT count(*) FROM {t}")
            print(f"  {t:15s}: {cur.fetchone()[0]}")
        cur.execute("SELECT count(*) FROM cell WHERE ranh_thua_id IS NOT NULL")
        print(f"  cell có geom    : {cur.fetchone()[0]}")
    print(f"\nXong. {n_geom} ô gắn geom (DCB02).")


if __name__ == "__main__":
    main()
